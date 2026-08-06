# -*- coding: utf-8 -*-
"""調査結果を Excel ブックとして書き出す。

成果物は xlsx であって PNG ではない。既存の調査結果ブック
(ExcelSQL/調査結果/*.xlsx) と同じ構成にする:

  調査結果シート … 期限超過ロットの一覧（46列のうち、いま出せる範囲）
  Sheet1        … 散布図の元データ。B列に通し番号、C〜H列に特性
  グラフ6枚     … Sheet1 に貼る。特性ごとに1枚、散布図

グラフの寸法は既存に合わせる。既存は高さ13行・幅2〜4列で並んでおり、標準の
列幅/行高で換算すると 1枚あたり約 7.6 x 6.5 cm、縦横比 1.17。matplotlib で
作っていた PNG は 1.67 で横長すぎた。
"""
import os
import re
from datetime import datetime
import sys

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

sys.stdout.reconfigure(encoding="utf-8")

# Sheet1 の C〜H 列。既存ブックの並びをそのまま踏襲する。
COLS = [("剥離表", "引きはがし強さ(表)"),
        ("剥離裏", "引きはがし強さ(裏)"),
        ("剥離処理後表", "引きはがし強さ　処理後(表)"),
        ("剥離処理後裏", "引きはがし強さ　処理後(裏)"),
        ("絶縁煮沸後", "絶縁抵抗(煮沸後)"),
        ("絶縁常態", "絶縁抵抗(常態)")]

RESULT_COLS = ["資材略称", "常非区分", "ケースNO", "製品ロット", "面", "品証ロット",
               "入荷日", "検査年月日", "保証期限", "製造品名", "製造年月日",
               "経過日数", "保証差日数",
               "剥離表", "剥離裏", "剥離処理後表", "剥離処理後裏",
               "絶縁常態", "絶縁煮沸後", "はんだ表", "はんだ裏"]

CHART_W, CHART_H = 7.6, 6.5      # cm。既存の実測値

# 帳票の作成時刻。実際の時刻を入れると、同じ調査から作った同じ帳票が毎回違う
# ファイルになる。日付そのものは調査結果の中（検査年月日・製造年月日）にあるので、
# 書庫の属性としては持たない。
FIXED_STAMP = datetime(2020, 1, 1)


def build(res, win, target_pos, path):
    wb = Workbook()

    ws = wb.active
    ws.title = "調査結果"
    ws.append(RESULT_COLS)
    for r in res["rows"]:
        ws.append([r.get(c, "") for c in RESULT_COLS])
    for i, c in enumerate(RESULT_COLS, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = max(10, len(c) + 4)
    ws.freeze_panes = "A2"

    s1 = wb.create_sheet("Sheet1")
    # 既存に合わせ、B列が通し番号、C〜H列が特性。1行目は空け、2行目を見出しにする。
    s1.cell(2, 2, 0)
    for j, (_, label) in enumerate(COLS, start=3):
        s1.cell(2, j, label)
    for i, row in enumerate(win, start=1):
        s1.cell(2 + i, 2, i)
        for j, (key, _) in enumerate(COLS, start=3):
            v = row.get(key)
            if isinstance(v, (int, float)):
                s1.cell(2 + i, j, v)
    n = len(win)

    # 対象ロットは、既存と同じく最終行に複製して別系列として描く
    if target_pos is not None:
        last = 3 + n
        s1.cell(last, 2, target_pos + 1)
        for j, (key, _) in enumerate(COLS, start=3):
            v = win[target_pos].get(key)
            if isinstance(v, (int, float)):
                s1.cell(last, j, v)

    for j, (key, label) in enumerate(COLS, start=3):
        ch = ScatterChart()
        ch.title = label
        ch.style = 13
        ch.x_axis.title = None
        ch.y_axis.title = None
        ch.height = CHART_H
        ch.width = CHART_W
        ch.legend = None

        xs = Reference(s1, min_col=2, min_row=3, max_row=2 + n)
        ys = Reference(s1, min_col=j, min_row=2, max_row=2 + n)
        se = Series(ys, xs, title_from_data=True)
        se.marker = Marker(symbol="circle", size=5)
        se.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
        ch.series.append(se)

        if target_pos is not None:
            tx = Reference(s1, min_col=2, min_row=3 + n, max_row=3 + n)
            ty = Reference(s1, min_col=j, min_row=3 + n, max_row=3 + n)
            se2 = Series(ty, tx, title="対象")
            se2.marker = Marker(symbol="square", size=10)
            se2.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            ch.series.append(se2)

        # 既存と同じく7行目から、3枚ずつ横に並べる
        idx = j - 3
        col = "BJR"[idx % 3]
        row = 7 + (idx // 3) * 14
        s1.add_chart(ch, "%s%d" % (col, row))

    # 作成時刻を書かせない。openpyxl は docProps/core.xml に保存した瞬間の時刻を
    # 入れるので、同じ調査から作った同じ帳票でも実行のたびにファイルが変わる。
    # 中身は一致しているのに「毎回違うもの」に見えてしまい、再現性を示せない。
    # 実測: 10回流して、この1か所だけが10通りに割れていた。
    wb.properties.created = FIXED_STAMP
    wb.properties.modified = FIXED_STAMP
    wb.properties.creator = ""
    wb.save(path)
    _freeze_stamps(path)
    return path


def _freeze_stamps(path):
    """保存後に、書庫の中の更新時刻を固定値へ書き直す。

    openpyxl は save() の中で modified を「いま」に上書きするので、事前に
    properties へ入れるだけでは効かない（実測: 10回流して 10 通りに割れ、差は
    docProps/core.xml の modified 1か所だけだった）。書き終えてから直す。

    書庫の各項目の日時も揃える。揃えないと、中身が同じでもファイルとしては
    毎回違うものになる。
    """
    import shutil
    import tempfile
    import zipfile

    stamp = (FIXED_STAMP.year, FIXED_STAMP.month, FIXED_STAMP.day, 0, 0, 0)
    iso = FIXED_STAMP.strftime("%Y-%m-%dT%H:%M:%SZ")
    with zipfile.ZipFile(path) as z:
        items = [(i.filename, z.read(i.filename)) for i in z.infolist()]

    fixed = []
    for name, data in items:
        if name == "docProps/core.xml":
            text = data.decode("utf-8")
            text = re.sub(r"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:)",
                          r"\g<1>%s\g<2>" % iso, text)
            data = text.encode("utf-8")
        fixed.append((name, data))

    tmp = tempfile.mktemp(suffix=".xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in fixed:
            info = zipfile.ZipInfo(name, date_time=stamp)
            info.compress_type = zipfile.ZIP_DEFLATED
            z.writestr(info, data)
    shutil.move(tmp, path)
