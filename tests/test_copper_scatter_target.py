"""散布図の対象ロットが真ん中に来ること、母集団が前後の実績であることを固定する。

PNG 側だけ、母集団に「期限超過63件」を並べ、その最後の点を対象として描いていた。
ブック側は最初から「対象を挟んだ前後」で描いていたので、同じ調査から出る2つの
成果物が食い違い、しかもスライドに載るのは間違っているほうだった。対象が右端に
来ると「周囲のどこに位置するか」というこの調査の主題がまるで読めない。

DB は要らない。母集団の作り方（window_around）と、描く位置の決め方だけを見る。
"""
import os
import sys

import pytest

SCRIPTS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "skills", "copper-foil-survey", "scripts")
sys.path.insert(0, SCRIPTS)


def _lots(n):
    return [{"ロットNO": "L%03d" % i, "LOT_Y": "24", "HINME": "X",
             "製造年月日": "24-01-%02d" % (i + 1), "剥離表": 0.5 + i * 0.001}
            for i in range(n)]


def test_target_sits_in_the_middle_when_there_is_room():
    from charts import WINDOW, window_around
    lots = _lots(200)
    win, pos = window_around(lots, "L100")
    assert len(win) == WINDOW
    # 真ん中。端に寄っていたら、この調査で見たいことが見えない
    assert abs(pos - WINDOW // 2) <= 1, pos
    assert win[pos]["ロットNO"] == "L100"


def test_target_at_the_start_still_gets_a_full_window():
    """対象が最初のロットでも、取れるだけ取って描くこと。"""
    from charts import WINDOW, window_around
    win, pos = window_around(_lots(200), "L000")
    assert len(win) == WINDOW
    assert win[pos]["ロットNO"] == "L000"
    assert pos == 0                      # 前が無いので端になるのは正しい


def test_target_at_the_end_still_gets_a_full_window():
    from charts import WINDOW, window_around
    win, pos = window_around(_lots(200), "L199")
    assert len(win) == WINDOW
    assert win[pos]["ロットNO"] == "L199"


def test_small_population_is_not_padded():
    """5ロットしか無い品名でも、そのまま5点で描くこと。"""
    from charts import window_around
    win, pos = window_around(_lots(5), "L002")
    assert len(win) == 5
    assert win[pos]["ロットNO"] == "L002"


def test_unknown_lot_falls_back_without_a_target():
    from charts import window_around
    win, pos = window_around(_lots(10), "存在しない")
    assert len(win) == 10
    assert pos is None                   # 対象が分からないなら印を付けない


def test_scatter_uses_the_window_not_the_expired_rows():
    """散布図が受け取るのは母集団であって、期限超過の一覧ではないこと。"""
    import inspect

    import survey
    sig = inspect.signature(survey._write_scatters)
    assert list(sig.parameters) == ["res", "win", "target_pos", "outdir"], sig
    src = inspect.getsource(survey._write_scatters)
    # 以前は res["rows"] を並べて最後の点を対象にしていた
    assert 'res["rows"]' not in src
    assert "xs[-1:]" not in src


def test_window_is_decided_once_and_shared():
    """母集団を2か所で別々に引かないこと。食い違いの元。"""
    import inspect

    import survey
    assert list(inspect.signature(survey._write_workbook).parameters) == \
        ["res", "win", "target_pos", "outdir"]
    book = inspect.getsource(survey._write_workbook)
    assert "fetch_neighbours" not in book, "ブック側でも引き直している"


def test_workbook_is_byte_identical_across_runs(tmp_path):
    """同じ入力から同じファイルが出ること。

    openpyxl は保存時に更新時刻を「いま」で書くので、事前に properties へ入れる
    だけでは効かない。10回流して10通りに割れ、差は docProps/core.xml の1か所
    だけだった。中身が同じなのに毎回違うファイルになると、再現性を示せない。
    """
    import hashlib

    import workbook

    res = {"material": "T", "rows": [{"資材略称": "T", "保証差日数": -128}]}
    win = [{"剥離表": 0.5 + i * 0.001, "絶縁常態": 1e13} for i in range(20)]

    seen = set()
    for i in range(3):
        p = tmp_path / ("b%d.xlsx" % i)
        workbook.build(res, win, 10, str(p))
        seen.add(hashlib.sha256(open(str(p), "rb").read()).hexdigest())
    assert len(seen) == 1, "実行ごとにファイルが変わる"


def test_workbook_still_opens_after_freezing_stamps(tmp_path):
    """時刻を固定したせいで壊れていないこと。"""
    from openpyxl import load_workbook

    import workbook

    res = {"material": "T", "rows": [{"資材略称": "T", "保証差日数": -1}]}
    win = [{"剥離表": 0.5 + i * 0.001} for i in range(12)]
    p = tmp_path / "b.xlsx"
    workbook.build(res, win, 6, str(p))
    wb = load_workbook(str(p))
    assert wb.sheetnames == ["調査結果", "Sheet1"]
    assert len(wb["Sheet1"]._charts) == 6
